#!/usr/bin/python3
# -*- coding: UTF-8 -*-
import csv
import os
from openpyxl import load_workbook

#读取登录用例数据并追加写入到csv文件
def xlsx_csv(xlsx_path, csv_path):
    PATH = os.path.abspath(os.path.dirname(os.getcwd()))
    old_file = PATH + csv_path
    try:
        os.remove(old_file)
    except Exception as msg:
        print('无旧的CSV文件，正常跳过此步骤',msg)
    wb = load_workbook(PATH+xlsx_path)
    ws = wb['Sheet1']

    for row in ws.iter_rows( min_row= None,max_row=None,min_col=None, max_col= None):
        line = [col.value for col in row]
        with open(PATH+csv_path,"a+",newline="",encoding='utf-8') as datacsv:
            csvwriter = csv.writer(datacsv,dialect = ("excel"))
            csvwriter.writerow(line)



if __name__ == '__main__':
    xlsx_csv("/data/case.xlsx","/data/csv/login.csv")

    pass
