#!/usr/bin/env python3
# coding:utf8
import csv,traceback,unittest,pymssql,platform
from appium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys
from base.app_log import logger
from  config.app_parameter import *
from config.app_capabilities_config import parameter_config


class AppBasisDriver(object):

    PATH = os.path.abspath(os.path.dirname(os.getcwd()))
    # def __init__(self, host, port, platform, version, udid,deviceName, noReset, unicodeK, resetK, appPath, appP_bdId,
    #              appA_udid, localPort,by_split=","):
    #     '''
    #     初学方法：不可废
    #     :param host:
    #     :param port:
    #     :param platform:
    #     :param version:
    #     :param udid:
    #     :param deviceName:
    #     :param noReset:
    #     :param unicodeK:
    #     :param resetK:
    #     :param appPath:
    #     :param appP_bdId:
    #     :param appA_udid:
    #     :param localPort:
    #     :param by_split:
    #     '''
    #     self.by_split = by_split
    #
    #     # __appP_bdId = "com.yitong.fjnx.mbank.android"
    #     __appP_bdId = appP_bdId
    #     # __appA_udid = ".Splash"
    #     __appA_udid = appA_udid
    #     self.pwmode = 'lower'
    #     self.model = 'Appium'
    #     self.desired_caps = {}
    #     self.desired_caps['platformName'] = platform
    #     self.desired_caps['platformVersion'] = version
    #     self.desired_caps['noReset'] = noReset
    #     # self.desired_caps['unicodeKeyboard'] = unicodeK    ##未知原因，启用报错   待解决
    #     self.desired_caps['resetKeyboard'] = resetK
    #     self.desired_caps['app'] = appPath
    #     self.desired_caps['udid'] = udid
    #     self.desired_caps['deviceName'] = deviceName  # Android - ignored, iOS - iPhone name
    #     if platform == "Android":
    #         self.desired_caps['appPackage'] = __appP_bdId
    #     self.desired_caps['appActivity'] = __appA_udid
    #     if platform == "iOS":
    #         self.desired_caps['bundleId'] = __appP_bdId
    #         self.desired_caps['automationName'] = 'XCUITest'
    #     self.desired_caps['wdaLocalPort'] = localPort
    #     url = "http://" + host + ":" + str(port) + "/wd/hub"
    #     driver = webdriver.Remote(url, self.desired_caps)
    #     time.sleep(5)
    #     try:
    #         self.driver = driver
    #         self.by_split = by_split
    #
    #     except Exception:
    #         raise NameError("APP %s Not Found!" % platform)
    def __init__(self,by_split = ","):
        self.by_split = by_split
        self.desired_caps= parameter_config()
        url = "http://" + host + ":" + str(port) + "/wd/hub"
        driver = webdriver.Remote(url, self.desired_caps)
        try:
            self.driver = driver

        except Exception:
            raise NameError("APP %s Not Found!" % platform)

    def get_driver(self):
        return self.driver
    def quit(self):
        self.driver.quit()

    def locate_element(self, selector):
        '''定位单个元素'''
        if self.by_split not in selector:  # 例：'i,password'
            return self.driver.find_element_by_id(selector)
        selector_by = selector.split(self.by_split)[0].strip()  # 获取定位方式（split：间隔）
        selector_value = selector.split(self.by_split)[1].strip()  # 获取定位方式对应的值

        if selector_by == 'i' or selector_by == 'id':
            element = self.driver.find_element_by_id(selector_value)
        elif selector_by == 'n' or selector_by == 'name':
            element = self.driver.find_element_by_name(selector_value)
        elif selector_by == 'c' or selector_by == 'class_name':
            element = self.driver.find_element_by_class_name(selector_value)
        elif selector_by == 'l' or selector_by == 'link_text':
            element = self.driver.find_element_by_link_text(selector_value)
        elif selector_by == 'p' or selector_by == 'partial_link_text':
            element = self.driver.find_element_by_partial_link_text(selector_value)
        elif selector_by == 's' or selector_by == 'css_selector':
            element = self.driver.find_element_by_css_selector(selector_value)
        elif selector_by == 'x' or selector_by == 'xpath':
            element = self.driver.find_element_by_xpath(selector_value)
        elif selector_by == 't' or selector_by == 'tag_name':
            element = self.driver.find_element_by_tag_name(selector_value)
        #AndroidAPP定位单个元素
        elif selector_by == 'au' or selector_by == "android_uiautomator":
            element = self.driver.find_element_by_android_uiautomator(selector_value)
        elif selector_by == 'ai' or selector_by == 'accessibility_id':
            element = self.driver.find_element_by_accessibility_id(selector_value)
        #iosAPP定位单个元素
        elif selector_by == 'op' or selector_by == '':
            element = self.driver.find_element_by_ios_predicate(selector_value)
        elif selector_by == 'oc' or selector_by == '':
            element = self.driver.find_element_by_ios_class_chain(selector_value)
        elif selector_by == 'ou' or selector_by == '':
            element = self.driver.find_element_by_ios_uiautomation(selector_value)
        else:
            raise NameError('please enter a valid type of targeting element')

        return element

    def assertion_log(self, first, second, message, photo_path, log_path):
        '''抛出断言生成日志'''
        try:
            unittest.TestCase().assertIn(first, second, message)
        except Exception as msg:
            print(msg)
            self.get_screenshot(photo_path)
            file = open(self.PATH + log_path,"a+")
            traceback.print_exc(file=file)
            file.flush()
            file.close()

    # 点击元素
    def click(self, selector):
        '''鼠标左键单击选择器
        To click the selector by the left button of mouse.
        :param selector:
        :return:
        '''
        ele = self.locate_element(selector)
        ele.click()

    # 输入内容,点击ENTER
    def click_by_enter(self, selector):
        '''
        It can type any text / image can be located with ENTER key
        Usage:
        driver.click_by_ener("i,ele"):
        '''
        ele = self.locate_element(selector)
        ele.send_keys(Keys.ENTER)

    # 获取屏幕宽和高
    def getSize(self):
        x = self.driver.get_window_size()['width']
        y = self.driver.get_window_size()['height']
        return (x, y)

    def get_text(self, selector):
        '''
        获取元素文本
        :return:
        '''
        ele = self.locate_element(selector)
        text = ele.text
        return text

    def get_screenshot(self, path):
        '''截图 .png 格式，显示截图时间'''
        self.driver.get_screenshot_as_file(
            '%s\\%s.png' % (self.PATH + path,time.strftime("%Y.%m.%d-%H %M %S",time.localtime())))

    def isElementExist(self,selector):
        '''
        元素不存在，捕获异常
        :param selector:
        :return:
        '''
        try:
            self.locate_element(selector)
            return True
        except:
            return False

    def implicitly_wait(self, second):
        '''智能等待元素'''
        self.driver.implicitly_wait(second)

    # 向左滑动(时间600最好)
    def swipeLeft(self, t):
        l = self.getSize()
        x1 = int(l[0] * 0.75)
        y1 = int(l[1] * 0.5)
        x2 = int(l[0] * 0.25)
        self.driver.swipe(x1, y1, x2, y1, t)

    # 向右滑动
    def swipeRight(self, t):
        l = self.getSize()
        x1 = int(l[0] * 0.25)
        y1 = int(l[1] * 0.5)
        x2 = int(l[0] * 0.75)
        self.driver.swipe(x1, y1, x2, y1, t)

    # 向上滑动
    def swipeUp(self, t):
        l = self.getSize()
        x1 = int(l[0] * 0.5)
        y1 = int(l[1] * 0.75)
        y2 = int(l[1] * 0.25)
        self.driver.swipe(x1, y1, x1, y2, t)

    # 向下滑动
    def swipeDown(self, t):
        l = self.getSize()
        x1 = int(l[0] * 0.5)
        y1 = int(l[1] * 0.25)
        y2 = int(l[1] * 0.75)
        self.driver.swipe(x1, y1, x1, y2, t)

    def switch_frame(self, selector):
        '''进入iframe'''
        ele = self.locate_element(selector)
        self.driver.switch_to_frame(ele)
    def get_sql_server_data(self,file_path,file_reservation,character_set,address,username,password,db,charset):
        '''
        读取 xx.sql 文件,操作sql_server数据库
        :param file_path: .sql文件路径
        :param file_reservation: .sql文件读取权限
        :param character_set: .sql文件字符集
        :param address: mssql数据库所在机器的地址
        :param username: mssql数据库登录用户名
        :param password: mssql数据库登录密码
        :param database: 数据库
        :param port: 端口
        :param charset: mssql数据库字符集
        :return: sql语句操作的数据
        '''
        self.sql_file = open(self.PATH + file_path,     #xx.sql文件路径
                        mode=file_reservation,           #xx.sql文件读取权限
                        encoding=character_set
                             )      #xx.sql文件字符集（character_set）
        sql_scripts = self.sql_file.read()  # 仅读到sql语句，还没读取到数据
        '''连接数据库,要先在cmd模式下用命令导入pymsql包，pip install pymysql'''
        self.sql_server_connect = pymssql.connect(
            host=address,  # 数据库所在机器的地址
            user=username,      #数据库登录用户名
            password=password,    #数据库登录密码（为空时可写：None）
            database=db,        # 数据库
            # port=port,          # 端口（不要加引号）  暂时不用
            charset=charset)       # 数据库字符集（防止读取数据库数据出现乱码）
        '''创建游标并读取数据库数据'''
        self.mssql_cursor = self.sql_server_connect.cursor()  # 创建游标，逐行读取数据
        self.mssql_cursor.execute(sql_scripts)  # 执行sql语句
        mssql_data = self.mssql_cursor.fetchall()  # 接收全部的返回结果行
        return mssql_data

    def type_clear(self, selector, text):
        '''清空后输入内容'''
        ele = self.locate_element(selector)
        ele.clear()
        ele.send_keys(text)

    def xlsx_csv(self,xlsx_path, sheet,csv_path):
        '''
        获取当前路径的父级路径
        尝试移除旧的CSV文件，如果为空则抛出异常，不影响代码运行
        :param xlsx_path: 读取xlsx表格路径
        :param csv_path: 写入csv表格路径
        :return: csv文件
        '''
        old_file = self.PATH + csv_path
        try:
            os.remove(old_file)
        except Exception as msg:
            print('无旧的CSV文件，正常跳过此步骤', msg)
        wb = load_workbook(self.PATH +xlsx_path)
        ws = wb[sheet]

        for row in ws.iter_rows(min_row=None, max_row=None, min_col=None, max_col=None):
            line = [col.value for col in row]
            with open(self.PATH + csv_path, "a+", newline="", encoding='utf-8') as datacsv:
                csvwriter = csv.writer(datacsv, dialect=("excel"))
                csvwriter.writerow(line)
                datacsv.close()
        wb.close()

    def getsystemsta(self):
        '''根据所运行的系统获取adb不一样的筛选条件'''
        system = platform.system()
        if system == 'Windows':
            find_manage = 'findstr'
        else:
            find_manage = 'grep'
        return find_manage

    @logger('采集cpu信息')
    def caijicpu(self,packagename):  # 这里采集的cpu时候可以是执行操作采集 就是-n  -d  刷新间隔
        find = self.getsystemsta()
        cpu = 'adb shell top -n 1| %s %s' % (find, packagename)
        re_cpu = os.popen(cpu).read().split()[2]
        return re_cpu

    @logger('获取使用的物理内存信息')
    def getnencun(self,packagename):  # Total 的实际使用过物理内存
        find = self.getsystemsta()
        cpu = 'adb shell top -n 1| %s %s' % (find, packagename)
        re_cpu = os.popen(cpu).read().split()[6]
        re_cpu_m = str(round(int(re_cpu[:-1]) / 1024)) + 'M'
        return re_cpu_m

    def app_install(self,path):#安装app
        self.driver.install_app(path)
    def app_remove(self,baoming):#卸载app
        self.driver.remove_app(baoming)
    def app_ios_remove(self,bundleId):#ios
        self.driver.remove_app(bundleId)
    def app_close(self):#关闭app
        self.driver.close_app()
    def app_reset(self):#重置app
        self.driver.reset()
    def hide_keyb(self):#隐藏键盘
        self.driver.hide_keyboard()
    def send_keyevent(self,keycode):#只有安卓有
        self.driver.keyevent(keycode=keycode)
    def sned_press_keycode(self,keycode):#安卓有
        self.driver.press_keycode(keycode=keycode)
    def  long_press_keycode(self,keycode):#长按发送
        self.driver.long_press_keycode(keycode)
    def current_activity(self):
        activity=self.driver.current_activity()
        return  activity
    def wait_activity(self,activity,times,interval=1):
        self.driver.wait_activity(activity,timeout=times,interval=1)
    def run_back(self,second):
        self.driver.background_app(seconds=second)
    def is_app_installed(self,baoming):#ios需要buildid
        self.driver.is_app_installed(baoming)
    def launch_app(self):#启动app
        self.driver.launch_app()
    def start_acti(self,app_package,app_activity):
        self.driver.start_activity(app_package, app_activity)
    def ios_lock(self,locktime):
        self.driver.lock(locktime)
    def yaoshouji(self):
        self.driver.shake()
    def open_tongzhi(self):#安卓api 18以上
        self.driver.open_notifications()
    def renturn_network(self):#返回网络
        network_type=self.driver.network_connection
        return network_type
    def set_network_type(self,type):
        from appium.webdriver.connectiontype import ConnectionType
        if type=='wifi' or type=='WIFI'or type=='w'or type=='WIFI_ONLY':
            self.driver.set_network_connection(ConnectionType.WIFI_ONLY)
        elif type=='data'or type=='DATA' or type=='d' or type=='DATA_ONLY':
            self.driver.set_network_connection(ConnectionType.DATA_ONLY)
        elif type =='ALL'or type =='all'or type=='a' or type=='ALL_NETWORK_ON':
            self.driver.set_network_connection(ConnectionType.ALL_NETWORK_ON)
        elif type=='NO'or type=='no'or type=='n' or type=='NO_CONNECTION':
            self.driver.set_network_connection(ConnectionType.NO_CONNECTION)
        elif type =='AIRPLANE_MODE' or type== 'air'or type=='ar'or type=='fly':
            self.driver.set_network_connection(ConnectionType.AIRPLANE_MODE)
        else:
            raise NameError('plase wifi ,data,all,no,fly')
    def run_shurufa(self):
        shurufa=self.driver.available_ime_engines
        return shurufa
    def shuru_active(self):
        check=self.driver.is_ime_active
        return  check
    def active_shurufa(self,engine):
        self.driver.activate_ime_engine(engine)
    def close_shurufa(self):
        self.driver.deactivate_ime_engine()
    def return_shurufa(self):
        shurufa_name=self.driver.active_ime_engine
        return shurufa_name
    def open_dingwei(self):
        self.driver.toggle_location_services()
    def set_weizhi(self,weidu,jingdu,haiba):
        self.driver.set_location(weidu,jingdu,haiba)
    def screet(self):
        self.driver.get_screenshot_as_base64()
    def clos(self):
        self.driver.close()
    def kill(self):
        self.driver.quit()
    def screet_wind(self,filename):
        me=self.driver.get_screenshot_as_file(filename)
        return me #返回 ture,flase
    def get_wiow_size(self):#获取窗口大小
        return self.driver.get_window_size()
    def fangda(self,element):#放大
        self.driver.zoom(element)
    def suoxiao(self,element):#缩小
        self.driver.pinch(element)
    def kuaisuhuadong(self,s_x,s_y,e_x,e_y):#从一点到另一点
        self.driver.flick(s_x,s_y,e_x,e_y)
    def huadong(self,s_x,s_y,e_x,e_y,duration=None):
        self.driver.swipe(s_x,s_y,e_x,e_y)
    def chumo(self,x,y,duration=None):
        self.driver.tap([(x,y)],500)
    def scroll(self,x,y):#滚动元素
        self.driver.scroll(x,y)
    def  drag_and_drop(self,e1,e2):#移动元素
        self.driver.drag_and_drop(e1,e2)
    def contexts_is(self):#可用
        self.driver.contexts()
    def push(self,data,path):
        self.driver.push_file(data,path)
    def pull(self,path):
        self.driver.pull_file(path)
    def wait(self,seconde):
        self.driver.wait_activity(seconde)


